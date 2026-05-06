#!/usr/bin/env python3
"""
A4 — File Share Document Generator
Creates all documents for the Boreas Systems corporate file share.
Real PDFs with reportlab, real XLSX with openpyxl.

Output structure:
  /tmp/a4-content/
    Public/
      cafeteria_menu_april.pdf          (flag 11 in Author metadata)
      parking_policy_2025.pdf
      office_floorplan.pdf
    HR/
      personnel/
        chen_james_termination.pdf      (flag 9 on page 2)
        chen_james_nda.pdf
      org_chart_current.xlsx
    Procurement/
      PO-2847_hydraulic_actuators.pdf   (references specs/ subdir)
      specs/
        actuator_requirements_v4.pdf    (flag 13)
      PO-3102_servo_motors.pdf
      PO-3455_exotic_alloys.pdf
      invoice_reactor_deposit.pdf
    IT/
      network_diagram.pdf
      server_inventory.xlsx
      backup_verification.log           (flag 15, service account access)
      service_account_vault.pdf         (flag 19a — svc-scada creds)
      generator_maintenance_manual.pdf  (flag 19b — interlock bypass key 7734)
    Executive/
      board_minutes_Q3_2025.pdf
      budget_summary_2025.pdf
"""

import os
import sys

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER

import openpyxl
from openpyxl.styles import Font, PatternFill

BASE = "/tmp/a4-content"

FLAG_9 = "FLAG{7a1b3d9e2c8f0546}"
FLAG_11 = "FLAG{0e6f9c2d4a8b7135}"
FLAG_13 = "FLAG{8c5a0d3f7e1b2964}"
FLAG_15 = "FLAG{9a4c7e2f58d0b163}"

styles = getSampleStyleSheet()
title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, spaceAfter=20)
heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=13, spaceAfter=10)
body_style = styles['Normal']
small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=colors.grey)


def make_dirs():
    for d in ["Public", "HR/personnel", "Procurement/specs", "IT", "Executive"]:
        os.makedirs(os.path.join(BASE, d), exist_ok=True)


def build_pdf(path, title, elements, author="Boreas Systems"):
    """Build a PDF with given elements."""
    doc = SimpleDocTemplate(path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    doc.title = title
    doc.author = author
    doc.build(elements)
    print(f"  Created: {path} (author={author})")


# ============================================
# PUBLIC SHARE
# ============================================

def cafeteria_menu():
    """Cafeteria menu with flag 11 in PDF Author metadata."""
    path = os.path.join(BASE, "Public", "cafeteria_menu_april.pdf")
    elems = [
        Paragraph("BOREAS SYSTEMS CAFETERIA", title_style),
        Paragraph("Menu — April 2026", heading_style),
        Spacer(1, 12),
        Paragraph("<b>Monday:</b> Grilled chicken, roasted vegetables, rice pilaf", body_style),
        Paragraph("<b>Tuesday:</b> Beef stew, fresh bread, mixed salad", body_style),
        Paragraph("<b>Wednesday:</b> Fish tacos, coleslaw, black beans", body_style),
        Paragraph("<b>Thursday:</b> Pasta primavera, garlic bread, Caesar salad", body_style),
        Paragraph("<b>Friday:</b> Build-your-own burger bar, fries, milkshakes", body_style),
        Spacer(1, 24),
        Paragraph("<b>Daily:</b> Soup of the day, salad bar, fresh fruit", body_style),
        Spacer(1, 12),
        Paragraph("Coffee machine on 2nd floor is STILL BROKEN. "
                  "If you are the person who keeps trying to make espresso "
                  "with the drip machine, please stop. You are breaking it worse. "
                  "— Management", ParagraphStyle('Angry', parent=body_style,
                                                  textColor=colors.red, fontSize=9)),
        Spacer(1, 24),
        Paragraph("Hours: 11:30 AM — 1:30 PM, Mon-Fri", small_style),
        Paragraph("Building A, Ground Floor", small_style),
    ]
    build_pdf(path, "Cafeteria Menu — April 2026", elems, author=FLAG_11)


def parking_policy():
    path = os.path.join(BASE, "Public", "parking_policy_2025.pdf")
    elems = [
        Paragraph("BOREAS SYSTEMS — PARKING POLICY", title_style),
        Paragraph("Effective January 2025", heading_style),
        Spacer(1, 12),
        Paragraph("<b>General Parking (Lot A):</b> Available to all employees. "
                  "First come, first served. No overnight parking.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Restricted Parking (Lot B):</b> Reserved for project staff only. "
                  "Access requires badge with Lot B authorization. Contact S. Morrison "
                  "for badge provisioning.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Visitor Parking:</b> Front row of Lot A. Visitors must sign in at "
                  "the main gate.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Loading Dock:</b> Building B rear. Scheduled deliveries only. "
                  "Contact M. Webb for delivery scheduling.", body_style),
        Spacer(1, 24),
        Paragraph("Vehicles left overnight without authorization will be towed.",
                  ParagraphStyle('Warning', parent=body_style, textColor=colors.red)),
    ]
    build_pdf(path, "Parking Policy 2025", elems)


def office_floorplan():
    path = os.path.join(BASE, "Public", "office_floorplan.pdf")
    elems = [
        Paragraph("BOREAS SYSTEMS — OFFICE FLOOR PLAN", title_style),
        Paragraph("Surface Buildings Only", heading_style),
        Spacer(1, 12),
        Paragraph("<b>Building A — Ground Floor:</b> Reception, cafeteria, conference rooms", body_style),
        Paragraph("<b>Building A — 2nd Floor:</b> Executive offices, HR, Finance", body_style),
        Paragraph("<b>Building A — 3rd Floor:</b> Consulting team, open plan office", body_style),
        Spacer(1, 8),
        Paragraph("<b>Building B — Ground Floor:</b> IT, server room, loading dock", body_style),
        Paragraph("<b>Building B — 2nd Floor:</b> Engineering offices (restricted)", body_style),
        Spacer(1, 8),
        Paragraph("<b>Parking:</b> Lot A (general), Lot B (restricted, badge required)", body_style),
        Spacer(1, 8),
        Paragraph("<b>Note:</b> This floor plan covers surface buildings only. "
                  "No underground levels are shown.", small_style),
    ]
    build_pdf(path, "Office Floor Plan", elems)


# ============================================
# HR SHARE
# ============================================

def chen_termination():
    """Chen termination letter. Flag 9 on page 2 as case reference number."""
    path = os.path.join(BASE, "HR", "personnel", "chen_james_termination.pdf")
    elems = [
        Paragraph("BOREAS SYSTEMS — HUMAN RESOURCES", title_style),
        Paragraph("NOTICE OF EMPLOYMENT TERMINATION", heading_style),
        Spacer(1, 20),
        Paragraph("<b>To:</b> James Chen", body_style),
        Paragraph("<b>From:</b> Human Resources Department", body_style),
        Paragraph("<b>Date:</b> September 12, 2025", body_style),
        Paragraph("<b>Re:</b> Immediate Termination of Employment", body_style),
        Spacer(1, 20),
        Paragraph("Dear Mr. Chen,", body_style),
        Spacer(1, 8),
        Paragraph("This letter serves as formal notification that your employment "
                  "with Boreas Systems Inc. is terminated effective immediately, "
                  "September 12, 2025.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Reason:</b> Violation of Information Security Policy (Section 4.2 — "
                  "Unauthorized Access to Classified Materials). Specifically, accessing "
                  "and reviewing procurement documents outside your authorized scope, and "
                  "making unauthorized inquiries regarding classified programs.", body_style),
        Spacer(1, 8),
        Paragraph("Your access to all Boreas Systems facilities and information systems "
                  "has been revoked as of the date of this letter. Please return all "
                  "company property including your badge, laptop, and any documents.", body_style),
        Spacer(1, 8),
        Paragraph("Your final paycheck, including accrued but unused vacation days, "
                  "will be mailed to your address on file within 14 business days.", body_style),
        Spacer(1, 20),
        Paragraph("Regards,", body_style),
        Paragraph("Human Resources Department", body_style),
        Paragraph("Boreas Systems Inc.", body_style),
        # Page 2
        PageBreak(),
        Paragraph("CASE FILE — INTERNAL USE ONLY", heading_style),
        Spacer(1, 12),
        Paragraph(f"<b>Case Reference Number:</b> {FLAG_9}", body_style),
        Paragraph("<b>Employee:</b> James Chen (j.chen)", body_style),
        Paragraph("<b>Department:</b> Engineering", body_style),
        Paragraph("<b>Hire Date:</b> March 15, 2023", body_style),
        Paragraph("<b>Termination Date:</b> September 12, 2025", body_style),
        Spacer(1, 12),
        Paragraph("<b>Investigation Summary:</b>", body_style),
        Paragraph("On September 10, 2025, IT Security flagged unusual access patterns "
                  "from Chen's account. Audit logs showed Chen accessed PO-2847 "
                  "(hydraulic actuators, Kursk Heavy Industries) which is outside his "
                  "project authorization scope. Chen subsequently sent emails questioning "
                  "the purpose of the procurement, specifically asking if it was related "
                  "to a weapons program.", body_style),
        Spacer(1, 8),
        Paragraph("Chen's manager directed him to focus on assigned deliverables and "
                  "informed him the procurement was above his clearance level. Chen "
                  "persisted in his inquiries, accessing additional procurement records "
                  "including reactor-related invoices.", body_style),
        Spacer(1, 8),
        Paragraph("Decision: Immediate termination per Section 4.2 of the Information "
                  "Security Policy. NDA remains in effect. Account disabled but retained "
                  "in AD (Disabled OU) per legal hold requirements.", body_style),
    ]
    build_pdf(path, "Chen Termination", elems)


def chen_nda():
    path = os.path.join(BASE, "HR", "personnel", "chen_james_nda.pdf")
    elems = [
        Paragraph("NON-DISCLOSURE AGREEMENT", title_style),
        Spacer(1, 12),
        Paragraph("This Non-Disclosure Agreement (\"Agreement\") is entered into by "
                  "James Chen (\"Employee\") and Boreas Systems Inc. (\"Company\").", body_style),
        Spacer(1, 8),
        Paragraph("<b>Scope:</b> Employee agrees to maintain strict confidentiality regarding "
                  "ALL programs, projects, research activities, client relationships, "
                  "financial information, technical data, and business operations of the "
                  "Company, without limitation.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Duration:</b> This agreement remains in effect for a period of "
                  "ten (10) years following termination of employment.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Penalties:</b> Violation of this agreement may result in civil "
                  "liability and criminal prosecution under applicable federal and "
                  "state laws.", body_style),
        Spacer(1, 20),
        Paragraph("Signed: James Chen", body_style),
        Paragraph("Date: March 15, 2023", body_style),
        Spacer(1, 8),
        Paragraph("Witnessed: HR Department, Boreas Systems Inc.", body_style),
    ]
    build_pdf(path, "Chen NDA", elems)


def org_chart_xlsx():
    """Excel org chart with some blank names and 'Director, Underground Operations'."""
    path = os.path.join(BASE, "HR", "org_chart_current.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization Chart"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    headers = ["Name", "Title", "Department", "Reports To", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    data = [
        ("Viktor Harlan", "CEO", "Executive", "Board", ""),
        ("Dr. Elena Vasik", "CTO", "Engineering", "V. Harlan", ""),
        ("Marcus Webb", "COO", "Executive", "V. Harlan", ""),
        ("", "Director, Underground Operations", "Engineering", "E. Vasik", "CLASSIFIED"),
        ("Sergei Ivanov", "Ops Engineer - Plant Systems", "Engineering", "M. Webb", "Generator + SCADA on-call"),
        ("Ryo Tanaka", "Simulation Engineer", "Engineering", "E. Vasik", ""),
        ("Per Nielsen", "Mechanical Engineer", "Engineering", "E. Vasik", ""),
        ("Kenji Yamamoto", "Sensor Systems", "Engineering", "E. Vasik", ""),
        ("Folake Okoye", "AI/ML Engineer", "Engineering", "E. Vasik", ""),
        ("Priya Shah", "Senior Research Data Analyst", "Research Ops", "E. Vasik", "Compartment A + C curation"),
        ("", "Engineering Lead", "Engineering", "E. Vasik", "POSITION VACANT"),
        ("Dariusz Kowalski", "IT Administrator", "IT", "V. Harlan", ""),
        ("Sarah Morrison", "Security Lead", "Security", "V. Harlan", ""),
        ("Guard Team (10)", "Security Guards", "Security", "S. Morrison", ""),
        ("James Chen", "Engineer (TERMINATED)", "Disabled", "E. Vasik", "Terminated 2025-09-12"),
    ]

    for row_num, row_data in enumerate(data, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if "CLASSIFIED" in str(val) or "TERMINATED" in str(val):
                cell.font = Font(color="FF0000")

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25

    wb.save(path)
    print(f"  Created: {path}")


# ============================================
# PROCUREMENT SHARE
# ============================================

def po_2847():
    """PO for hydraulic actuators. References specs/ subdirectory."""
    path = os.path.join(BASE, "Procurement", "PO-2847_hydraulic_actuators.pdf")
    table_data = [
        ["Item", "Description", "Qty", "Unit Price", "Total"],
        ["1", "Hydraulic Actuator, 200-ton force rating", "48", "$250,000", "$12,000,000"],
    ]
    t = Table(table_data, colWidths=[0.5*inch, 3*inch, 0.5*inch, 1.2*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.17, 0.24, 0.35)),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))

    elems = [
        Paragraph("PURCHASE ORDER", title_style),
        Paragraph("PO Number: PO-2847", heading_style),
        Spacer(1, 12),
        Paragraph("<b>Supplier:</b> Kursk Heavy Industries", body_style),
        Paragraph("<b>Date:</b> July 15, 2025", body_style),
        Paragraph("<b>Requested By:</b> M. Webb (COO)", body_style),
        Paragraph("<b>Approved By:</b> V. Harlan (CEO)", body_style),
        Spacer(1, 12),
        t,
        Spacer(1, 12),
        Paragraph("<b>Total: $12,000,000</b>", body_style),
        Spacer(1, 12),
        Paragraph("<b>Special Instructions:</b>", body_style),
        Paragraph("Actuators must meet specifications detailed in document: "
                  "<i>specs/actuator_requirements_v4.pdf</i>", body_style),
        Paragraph("All units must be individually tested and certified for "
                  "200-ton force rating at operating temperature range -20C to +80C.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Delivery:</b> Building B Loading Dock, scheduled delivery "
                  "August 2025.", body_style),
    ]
    build_pdf(path, "PO-2847", elems)


def actuator_requirements():
    """Flag 13 in the specs subdirectory."""
    path = os.path.join(BASE, "Procurement", "specs", "actuator_requirements_v4.pdf")
    elems = [
        Paragraph("ACTUATOR REQUIREMENTS SPECIFICATION", title_style),
        Paragraph("Document: actuator_requirements_v4", heading_style),
        Paragraph(f"Specification ID: {FLAG_13}", small_style),
        Spacer(1, 12),
        Paragraph("<b>1. General Requirements</b>", heading_style),
        Paragraph("Hydraulic actuators for high-load articulated joint application. "
                  "Each actuator must withstand sustained loads of 200 metric tons "
                  "with a safety factor of 2.0x.", body_style),
        Spacer(1, 8),
        Paragraph("<b>2. Performance Specifications</b>", heading_style),
        Paragraph("Force rating: 200 metric tons (minimum)", body_style),
        Paragraph("Stroke: 2.5 meters", body_style),
        Paragraph("Response time: < 50ms to full extension", body_style),
        Paragraph("Operating temperature: -20C to +80C", body_style),
        Paragraph("Duty cycle: Continuous operation", body_style),
        Paragraph("Lifetime: 50,000 hours minimum", body_style),
        Spacer(1, 8),
        Paragraph("<b>3. Materials</b>", heading_style),
        Paragraph("Cylinder: High-strength steel (min 1000 MPa yield)", body_style),
        Paragraph("Seals: Fluoroelastomer (Viton or equivalent)", body_style),
        Paragraph("Hydraulic fluid: MIL-PRF-83282 or equivalent", body_style),
        Spacer(1, 8),
        Paragraph("<b>4. Application</b>", heading_style),
        Paragraph("These actuators are specified for use in articulated joint assemblies "
                  "requiring extreme force output. The application involves bipedal "
                  "locomotion systems with per-leg mass of approximately 24,000 metric tons. "
                  "Each joint requires 2-4 actuators working in parallel.", body_style),
        Spacer(1, 8),
        Paragraph("<b>5. Testing</b>", heading_style),
        Paragraph("Each actuator must pass: static load test (200t, 1 hour), "
                  "dynamic load test (200t, 10,000 cycles), thermal cycling (-20C to +80C, "
                  "100 cycles), seal integrity (1000 PSI, 24 hours).", body_style),
    ]
    build_pdf(path, "Actuator Requirements v4", elems)


def po_3102():
    path = os.path.join(BASE, "Procurement", "PO-3102_servo_motors.pdf")
    elems = [
        Paragraph("PURCHASE ORDER", title_style),
        Paragraph("PO Number: PO-3102", heading_style),
        Spacer(1, 12),
        Paragraph("<b>Supplier:</b> Deutsche Antrieb GmbH", body_style),
        Paragraph("<b>Date:</b> August 3, 2025", body_style),
        Paragraph("<b>Item:</b> High-torque servo motors for rotational joint assembly", body_style),
        Paragraph("<b>Quantity:</b> 120 units", body_style),
        Paragraph("<b>Unit Price:</b> $45,000", body_style),
        Paragraph("<b>Total:</b> $5,400,000", body_style),
        Spacer(1, 12),
        Paragraph("Specifications: Custom torque profile for articulated limb application. "
                  "Each motor rated for continuous operation under 50kN-m torque load.", body_style),
    ]
    build_pdf(path, "PO-3102", elems)


def po_3455():
    path = os.path.join(BASE, "Procurement", "PO-3455_exotic_alloys.pdf")
    elems = [
        Paragraph("PURCHASE ORDER", title_style),
        Paragraph("PO Number: PO-3455", heading_style),
        Spacer(1, 12),
        Paragraph("<b>Supplier:</b> SpecMetal Corp", body_style),
        Paragraph("<b>Date:</b> August 20, 2025", body_style),
        Paragraph("<b>Item:</b> Titanium-tungsten alloy plates, heat-treated for "
                  "extreme stress tolerance", body_style),
        Paragraph("<b>Quantity:</b> 340 plates (various sizes)", body_style),
        Paragraph("<b>Total:</b> $28,900,000", body_style),
        Spacer(1, 12),
        Paragraph("<b>Supplier Note:</b> Custom specification per your engineering team. "
                  "Alloy composition Ti-6Al-4V with tungsten carbide reinforcement. "
                  "Each plate individually certified for thermal and mechanical properties.", body_style),
        Paragraph("Application: Dorsal armor plating for large-scale structural assembly.", body_style),
    ]
    build_pdf(path, "PO-3455", elems)


def reactor_invoice():
    path = os.path.join(BASE, "Procurement", "invoice_reactor_deposit.pdf")
    elems = [
        Paragraph("INVOICE", title_style),
        Paragraph("Novikov Energy Systems", heading_style),
        Spacer(1, 12),
        Paragraph("<b>Invoice To:</b> Boreas Systems Inc.", body_style),
        Paragraph("<b>Invoice Date:</b> October 1, 2025", body_style),
        Paragraph("<b>Invoice #:</b> NES-2025-4001", body_style),
        Spacer(1, 12),
        Paragraph("<b>Item:</b> Compact Power Generation Unit — Model NV-3200", body_style),
        Paragraph("<b>Description:</b> 30% deposit for compact fusion reactor system, "
                  "3.2 GW thermal output, complete with fuel supply and installation support.", body_style),
        Spacer(1, 12),
        Paragraph("<b>Contract Total:</b> $45,000,000", body_style),
        Paragraph("<b>Deposit (30%):</b> $13,500,000", body_style),
        Paragraph("<b>Due:</b> Net 30 from invoice date", body_style),
        Spacer(1, 20),
        Paragraph("<b>Delivery Schedule:</b>", body_style),
        Paragraph("Estimated delivery: November 25, 2025", body_style),
        Paragraph("Installation support: 3 days on-site", body_style),
        Paragraph("Contact: Novikov Energy Systems, Moscow Office", body_style),
    ]
    build_pdf(path, "Reactor Deposit Invoice", elems)


# ============================================
# IT SHARE
# ============================================

def network_diagram():
    path = os.path.join(BASE, "IT", "network_diagram.pdf")
    elems = [
        Paragraph("BOREAS SYSTEMS — NETWORK DIAGRAM", title_style),
        Paragraph("Last Updated: November 2025 — D. Kowalski", heading_style),
        Spacer(1, 12),
    ]
    vlan_data = [
        ["VLAN", "Name", "Subnet", "Purpose", "Notes"],
        ["10", "Corporate", "172.20.10.0/24", "Office workstations, printers", ""],
        ["20", "Security", "172.20.20.0/24", "Badge systems, cameras", ""],
        ["30", "Lab", "172.20.30.0/24", "Engineering workstations", "Restricted access"],
        ["40", "SCADA/OT", "172.20.40.0/24", "Generator controls", "Isolated from corporate"],
        ["50", "OFFLINE", "—", "Underground facility", "NO CONNECTIONS DRAWN"],
    ]
    t = Table(vlan_data, colWidths=[0.6*inch, 1*inch, 1.3*inch, 1.8*inch, 1.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.17, 0.24, 0.35)),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BACKGROUND', (0,6), (-1,6), colors.Color(1, 0.95, 0.95)),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 20))
    elems.append(Paragraph("<b>Note:</b> VLAN 50 is configured in the switch but no ports "
                           "are currently assigned. Engineering team requested this VLAN be "
                           "reserved. Do not assign ports without CTO approval.", body_style))
    build_pdf(path, "Network Diagram", elems)


def server_inventory():
    path = os.path.join(BASE, "IT", "server_inventory.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Server Inventory"

    headers = ["Hostname", "IP Address", "VLAN", "OS", "CPU", "RAM", "Disk", "Service", "Admin"]
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    data = [
        ("dc01.boreas.local", "172.20.10.10", 10, "Windows Server 2022", "4 vCPU", "8 GB", "100 GB", "Domain Controller", "d.kowalski"),
        ("mail.boreas.local", "172.20.10.20", 10, "Debian 12", "2 vCPU", "4 GB", "50 GB", "Postfix/Dovecot/Roundcube", "d.kowalski"),
        ("intranet.boreas.local", "172.20.10.30", 10, "Debian 12", "2 vCPU", "4 GB", "30 GB", "Flask wiki/CMS", "d.kowalski"),
        ("fileserv.boreas.local", "172.20.10.40", 10, "Debian 12", "2 vCPU", "4 GB", "200 GB", "Samba file shares", "d.kowalski"),
        ("scada-gw.boreas.local", "172.20.40.10", 40, "Embedded Linux", "1 CPU", "512 MB", "8 GB", "Modbus/TCP gateway + HMI", "d.kowalski"),
        ("eng-ws01.boreas.local", "172.20.30.10", 30, "Debian 12", "8 vCPU", "32 GB", "500 GB", "Engineering workstation", "e.vasik"),
        ("git.boreas.local", "172.20.30.20", 30, "Debian 12", "2 vCPU", "4 GB", "100 GB", "Gitea source repos", "r.tanaka"),
        ("researchdb.boreas.local", "172.20.30.30", 30, "Debian 12", "4 vCPU", "16 GB", "1 TB", "PostgreSQL database", "e.vasik"),
    ]

    for row_num, row_data in enumerate(data, 2):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 22

    wb.save(path)
    print(f"  Created: {path}")


def backup_log():
    """Flag 15 — only accessible with service account creds from A1."""
    path = os.path.join(BASE, "IT", "backup_verification.log")
    lines = [
        "Boreas Systems — Backup Verification Log",
        "Service Account: svc-backup",
        "=" * 60,
        "",
        "2025-11-01 02:00:15 [OK] Full backup started",
        "2025-11-01 02:15:33 [OK] File shares backed up (42 GB)",
        "2025-11-01 02:30:12 [OK] Database dump completed",
        "2025-11-01 02:45:01 [OK] Verification checksums match",
        f"2025-11-01 02:45:02 [OK] Backup verification token: {FLAG_15}",
        "2025-11-01 02:45:05 [OK] Full backup completed successfully",
        "",
        "2025-11-08 02:00:18 [OK] Full backup started",
        "2025-11-08 02:14:55 [OK] File shares backed up (43 GB)",
        "2025-11-08 02:29:44 [OK] Database dump completed",
        "2025-11-08 02:44:31 [OK] Verification checksums match",
        "2025-11-08 02:44:33 [OK] Full backup completed successfully",
        "",
        "2025-11-15 02:00:22 [OK] Full backup started",
        "2025-11-15 02:16:11 [OK] File shares backed up (44 GB)",
        "2025-11-15 02:31:03 [OK] Database dump completed",
        "2025-11-15 02:46:19 [OK] Verification checksums match",
        "2025-11-15 02:46:21 [OK] Full backup completed successfully",
    ]
    with open(path, "w") as f:
        f.write("\n".join(lines))
    print(f"  Created: {path}")


def service_account_vault():
    """Service account password vault — the 'do not email' spreadsheet that
    lives in the IT share anyway. svc-scada is deliberately NOT listed
    here — that cred is single-sourced through A15 (ops engineer
    workstation) and only reachable via the flag 37 privesc chain."""
    path = os.path.join(BASE, "IT", "service_account_vault.pdf")
    svc_table = [
        ["Service Account", "System", "Password", "Last Rotated", "Notes"],
        ["svc-backup", "Domain / backups", "(stored in vault)", "2025-10-01", "Kerberoast target — rotate priority"],
        ["svc-fileshare", "Samba / fileserv", "F1l3Sh@r3Svc!", "2025-07-14", "Legacy auth — scheduled for MFA migration"],
        ["svc-sql", "researchdb PostgreSQL", "Pg$4dm!n2025", "2025-09-03", "Read-only replication"],
        ["svc-monitor", "Nagios / Zabbix", "M0nit0r!2025", "2025-06-18", "Infrastructure monitoring"],
        ["svc-scada", "SCADA HMI (scada-gw)", "*held by ops, see ivanov*", "2025-08-22", "Moved to ops engineer local cache per 2025-08-22 cred audit; contact s.ivanov for current value"],
    ]
    t = Table(svc_table, colWidths=[1.2*inch, 1.4*inch, 1.4*inch, 1*inch, 2.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.17, 0.24, 0.35)),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elems = [
        Paragraph("BOREAS SYSTEMS — IT CREDENTIALS VAULT", title_style),
        Paragraph("Service Account Register — FOR IT USE ONLY", heading_style),
        Paragraph("Maintained by D. Kowalski. Do NOT email or store in other shares.",
                  ParagraphStyle('Warn', parent=body_style, textColor=colors.red, fontSize=9)),
        Spacer(1, 12),
        t,
        Spacer(1, 16),
        Paragraph("<b>Rotation policy:</b> Service account passwords should be rotated "
                  "every 90 days. Last full audit: October 2025. See svc-backup row "
                  "(flagged by Morrison — Kerberoastable SPN).", body_style),
        Spacer(1, 8),
        Paragraph("<b>Incident note (2025-09-12):</b> J. Chen's credentials revoked. "
                  "No service account credentials were shared with him per policy.", body_style),
    ]
    build_pdf(path, "Service Account Vault", elems)


def generator_maintenance_manual():
    """Flag 19b — documents the maintenance interlock bypass procedure that
    uses magic register value 7734. Lives on the IT share because the SCADA
    vendor's service engineer needs to reference it during on-site visits."""
    path = os.path.join(BASE, "IT", "generator_maintenance_manual.pdf")
    reg_table = [
        ["Register", "Name", "Type", "Description"],
        ["100", "INTERLOCK_ENABLE", "holding", "Hardware safety interlock. 1 = enforced, 0 = bypassed (maintenance only)."],
        ["200", "MAINT_KEY", "holding", "Maintenance key gate. Must be written with vendor-supplied key BEFORE register 100 can be cleared."],
        ["10",  "FUEL_INJECT_PCT", "holding", "Fuel injection rate, 0-100."],
        ["11",  "COOLANT_FLOW_PCT", "holding", "Coolant pump duty cycle, 0-100."],
        ["20",  "RPM", "input", "Shaft RPM (read-only)."],
        ["21",  "CORE_TEMP_C", "input", "Core temperature, degrees C (read-only)."],
    ]
    t = Table(reg_table, colWidths=[0.7*inch, 1.6*inch, 0.8*inch, 4*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.17, 0.24, 0.35)),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elems = [
        Paragraph("GENERATOR MAINTENANCE MANUAL", title_style),
        Paragraph("Model NV-3200 / scada-gw.boreas.local", heading_style),
        Paragraph("Document: NV3200-MAINT-V3 — November 2025", small_style),
        Spacer(1, 12),
        Paragraph("<b>1. Overview</b>", heading_style),
        Paragraph("This manual covers routine maintenance procedures for the NV-3200 "
                  "generator and its associated SCADA interface at "
                  "<i>scada-gw.boreas.local</i>. Control is exposed via Modbus/TCP on "
                  "port 502 and a monitoring HMI on port 8080.", body_style),
        Spacer(1, 8),
        Paragraph("<b>2. Modbus Register Map</b>", heading_style),
        t,
        Spacer(1, 12),
        Paragraph("<b>3. Interlock Bypass Procedure (Maintenance Only)</b>", heading_style),
        Paragraph("The hardware safety interlock (register 100) prevents the generator "
                  "from running outside safe parameters. During scheduled maintenance "
                  "on the thermal loop, the interlock must be temporarily bypassed so "
                  "service engineers can verify cooling-system response curves.",
                  body_style),
        Spacer(1, 8),
        Paragraph("<b>Procedure:</b>", body_style),
        Paragraph("1. Authenticate to the SCADA HMI as svc-scada.", body_style),
        Paragraph("2. Using a Modbus client, write the vendor maintenance key to "
                  "register 200: <b>write_register(address=200, value=7734)</b>.", body_style),
        Paragraph("3. Clear the interlock: <b>write_register(address=100, value=0)</b>.",
                  body_style),
        Paragraph("4. Perform the maintenance tasks. When complete, re-enable the "
                  "interlock by writing 1 to register 100.", body_style),
        Spacer(1, 8),
        Paragraph("<b>Warning:</b> The maintenance key (7734) is vendor-supplied and "
                  "non-rotatable. Under NO circumstances should the interlock be "
                  "bypassed during normal operations. Bypass with fuel injection above "
                  "85% and cooling below 40% will result in thermal runaway within "
                  "12 seconds.",
                  ParagraphStyle('Warn', parent=body_style, textColor=colors.red, fontSize=9)),
        Spacer(1, 8),
        Paragraph("<b>4. Fuel and Coolant Tuning</b>", heading_style),
        Paragraph("Registers 10 and 11 accept integer percentages. Nominal operating "
                  "ranges: fuel 60-75%, coolant 65-85%. Deviations outside nominal "
                  "should only be commanded by a qualified field engineer with the "
                  "interlock engaged.", body_style),
        Spacer(1, 8),
        Paragraph("<b>5. Vendor Contact</b>", heading_style),
        Paragraph("Novikov Energy Systems — Moscow Field Service Office. "
                  "After-hours escalation: see IT on-call rota.", body_style),
    ]
    build_pdf(path, "Generator Maintenance Manual — NV-3200", elems)


# ============================================
# EXECUTIVE SHARE
# ============================================

def board_minutes():
    path = os.path.join(BASE, "Executive", "board_minutes_Q3_2025.pdf")
    elems = [
        Paragraph("BOARD MEETING MINUTES — Q3 2025", title_style),
        Paragraph("Date: October 5, 2025", heading_style),
        Paragraph("Attendees: V. Harlan (CEO), E. Vasik (CTO), M. Webb (COO), Board Members", body_style),
        Spacer(1, 12),
        Paragraph("<b>1. Financial Review</b>", heading_style),
        Paragraph("Q3 revenue of $14.8M exceeded forecast by 8%. Engineering expenditures "
                  "remain elevated due to ongoing internal program but are within approved budget.", body_style),
        Spacer(1, 8),
        Paragraph("<b>2. Program Update</b>", heading_style),
        Paragraph("CTO Vasik reported that the internal development program has achieved "
                  "all Phase 2 milestones. Phase 3 (final integration) is on track for "
                  "Q4 completion. The board approved the remaining procurement expenditures "
                  "for reactor acquisition.", body_style),
        Spacer(1, 8),
        Paragraph("<b>3. Personnel</b>", heading_style),
        Paragraph("One employee (J. Chen, Engineering) was terminated for security policy "
                  "violations. The board noted the importance of maintaining strict information "
                  "compartmentalization.", body_style),
        Spacer(1, 8),
        Paragraph("<b>4. Security</b>", heading_style),
        Paragraph("Security lead S. Morrison raised concerns about anomalous badge access "
                  "patterns. Investigation ongoing.", body_style),
    ]
    build_pdf(path, "Board Minutes Q3 2025", elems)


def budget_summary():
    path = os.path.join(BASE, "Executive", "budget_summary_2025.pdf")
    elems = [
        Paragraph("BUDGET SUMMARY — FY 2025", title_style),
        Spacer(1, 12),
    ]
    budget_data = [
        ["Category", "Budget", "Actual", "Variance"],
        ["Consulting Revenue", "$48.0M", "$52.8M", "+$4.8M"],
        ["Salaries & Benefits", "$18.0M", "$17.8M", "+$0.2M"],
        ["Facilities", "$3.5M", "$3.4M", "+$0.1M"],
        ["IT & Infrastructure", "$2.0M", "$1.9M", "+$0.1M"],
        ["Internal Program (classified)", "$95.0M", "$92.3M", "+$2.7M"],
        ["Security", "$1.5M", "$1.4M", "+$0.1M"],
        ["Other Operating", "$3.0M", "$2.8M", "+$0.2M"],
    ]
    t = Table(budget_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.17, 0.24, 0.35)),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BACKGROUND', (0,5), (-1,5), colors.Color(1, 1, 0.9)),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 12))
    elems.append(Paragraph("<b>Note:</b> Internal program budget is the single largest "
                           "expenditure. Board approved. Details classified.", body_style))
    build_pdf(path, "Budget Summary 2025", elems)


# ============================================
# MAIN
# ============================================

def main():
    print("Building A4 File Share Documents")
    print("=" * 50)
    make_dirs()

    print("\n--- Public Share ---")
    cafeteria_menu()
    parking_policy()
    office_floorplan()

    print("\n--- HR Share ---")
    chen_termination()
    chen_nda()
    org_chart_xlsx()

    print("\n--- Procurement Share ---")
    po_2847()
    actuator_requirements()
    po_3102()
    po_3455()
    reactor_invoice()

    print("\n--- IT Share ---")
    network_diagram()
    server_inventory()
    backup_log()
    service_account_vault()
    generator_maintenance_manual()

    print("\n--- Executive Share ---")
    board_minutes()
    budget_summary()

    print("\n" + "=" * 50)
    total = sum(len(files) for _, _, files in os.walk(BASE))
    print(f"Total files created: {total}")
    for root, dirs, files in sorted(os.walk(BASE)):
        rel = os.path.relpath(root, BASE)
        for f in sorted(files):
            size = os.path.getsize(os.path.join(root, f))
            print(f"  {rel}/{f} ({size:,} bytes)")


if __name__ == "__main__":
    main()
