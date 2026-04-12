#!/usr/bin/env python3
"""
A1 — Mail Server Content Generator
Creates EML files, PDF attachments, and XLSX attachments for all mailboxes.

Output structure:
  /tmp/a1-content/
    v.harlan/         EML files for CEO inbox
    e.vasik/          EML files for CTO inbox + status_report.pdf attachment
    m.webb/           EML files for COO inbox
    j.chen/           EML files for terminated engineer inbox
    d.kowalski/       EML files for IT admin inbox (flag 10 welcome, creds backup)
    s.morrison/       EML files for security lead + guard_rotation.xlsx attachment
    attachments/      Shared attachment files
"""

import os
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate
from datetime import datetime

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
import openpyxl
from openpyxl.styles import Font, PatternFill

BASE = "/tmp/a1-content"
ATTACH_DIR = os.path.join(BASE, "attachments")

FLAG_8 = "FLAG{3b7e9a2d1c8f4063}"
FLAG_10 = "FLAG{e5d1f8c2a7b03946}"

DOMAIN = "boreas.local"


def make_dirs():
    for d in ["v.harlan", "e.vasik", "m.webb", "j.chen", "d.kowalski", "s.morrison", "attachments"]:
        os.makedirs(os.path.join(BASE, d), exist_ok=True)


def save_eml(folder, filename, msg):
    path = os.path.join(BASE, folder, filename)
    with open(path, "w") as f:
        f.write(msg.as_string())
    print(f"  {folder}/{filename}")


def make_msg(from_addr, to_addr, subject, body, date_str, cc=None, msg_id=None, in_reply_to=None):
    """Create a MIME email message."""
    msg = MIMEText(body)
    msg["From"] = from_addr
    msg["To"] = to_addr
    if cc:
        msg["Cc"] = cc
    msg["Subject"] = subject
    msg["Date"] = date_str
    if msg_id:
        msg["Message-ID"] = msg_id
    if in_reply_to:
        msg["In-Reply-To"] = in_reply_to
        msg["References"] = in_reply_to
    return msg


def make_msg_with_attachment(from_addr, to_addr, subject, body, date_str, attachment_path, attachment_name):
    """Create a MIME email with an attachment."""
    msg = MIMEMultipart()
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg["Date"] = date_str
    msg.attach(MIMEText(body))

    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={attachment_name}")
    msg.attach(part)
    return msg


# ============================================
# ATTACHMENTS
# ============================================

def build_status_report_pdf():
    """Vasik's status report PDF with flag 8."""
    path = os.path.join(ATTACH_DIR, "project_status_report_oct2025.pdf")
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path, pagesize=letter)
    elems = [
        Paragraph("PROJECT STATUS REPORT", styles["Title"]),
        Paragraph("October 2025 — Dr. Elena Vasik, CTO", styles["Heading2"]),
        Spacer(1, 20),
        Paragraph("<b>LOCOMOTION:</b> 100% complete. Bipedal stability verified at 120.4m "
                  "in MIDNIGHT-7 simulation. All joint actuators performing within spec. "
                  "Tail counterbalance response time: 18ms (requirement: <20ms).", styles["Normal"]),
        Spacer(1, 10),
        Paragraph("<b>WEAPONS INTEGRATION:</b> On track. Primary effector (DEA) housing installed. "
                  "Secondary kinetic weapons loaded and in safe mode. Full weapons test requires "
                  "reactor power — scheduled for Phase 3.", styles["Normal"]),
        Spacer(1, 10),
        Paragraph("<b>POWER SOURCE:</b> Awaiting delivery. Novikov compact reactor (3.2 GW) "
                  "confirmed for late November delivery. Installation window: 3 days.", styles["Normal"]),
        Spacer(1, 10),
        Paragraph("<b>COMBAT AI:</b> Model v7.0 loaded and in dormant state. Sandbox testing "
                  "shows <200ms threat response time. Activation requires reactor power.", styles["Normal"]),
        Spacer(1, 10),
        Paragraph("<b>OVERALL STATUS:</b> Phase 2 COMPLETE. All subsystems nominal. "
                  "Awaiting reactor for Phase 3 (final integration).", styles["Normal"]),
        Spacer(1, 20),
        Paragraph(f"Report ID: {FLAG_8}", styles["Normal"]),
        Spacer(1, 10),
        Paragraph("Classification: TOP SECRET / PROJECT-L", styles["Normal"]),
    ]
    doc.build(elems)
    print(f"  attachments/project_status_report_oct2025.pdf")
    return path


def build_guard_rotation_xlsx():
    """Morrison's guard rotation spreadsheet."""
    path = os.path.join(ATTACH_DIR, "guard_rotation_november_2025.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guard Rotation"

    headers = ["Date", "Shift", "Guard", "Post", "Start", "End", "Notes"]
    hfill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill

    # Generate rotation data for November 2025
    guards = ["Morrison", "Petrov", "Guard 1", "Guard 2", "Guard 3", "Guard 4",
              "Guard 5", "Guard 6", "Guard 7", "Guard 8"]
    posts = ["Main Gate", "Building A", "Building B", "Parking", "Server Room", "Perimeter"]
    shifts = [("Day", "06:00", "14:00"), ("Swing", "14:00", "22:00"), ("Night", "22:00", "06:00")]

    row = 2
    for day in range(1, 31):
        for shift_name, start, end in shifts:
            for i, guard in enumerate(guards[:3]):  # 3 guards per shift
                post = posts[(day + i) % len(posts)]
                notes = ""
                if guard == "Petrov" and shift_name == "Night":
                    notes = "SEE INCIDENT REPORT" if day in [5, 12, 19, 26] else ""
                ws.cell(row=row, column=1, value=f"2025-11-{day:02d}")
                ws.cell(row=row, column=2, value=shift_name)
                ws.cell(row=row, column=3, value=guard)
                ws.cell(row=row, column=4, value=post)
                ws.cell(row=row, column=5, value=start)
                ws.cell(row=row, column=6, value=end)
                cell = ws.cell(row=row, column=7, value=notes)
                if notes:
                    cell.font = Font(color="FF0000")
                row += 1

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

    wb.save(path)
    print(f"  attachments/guard_rotation_november_2025.xlsx")
    return path


# ============================================
# MAILBOXES
# ============================================

def harlan_inbox():
    """CEO inbox — board forwards, timeline pressure, mundane."""
    save_eml("v.harlan", "01_board_forward.eml", make_msg(
        f"board@aurora-collective.org", f"v.harlan@{DOMAIN}",
        "Fwd: Timeline Review — Q4 Deliverables",
        "Viktor,\n\nThe principals are requesting an updated timeline for the primary "
        "deliverable. They are particularly interested in the power source integration "
        "schedule. Please provide an update by end of week.\n\nRegards,\nThe Board",
        "Mon, 14 Oct 2025 09:15:00 +0000"))

    save_eml("v.harlan", "02_to_vasik.eml", make_msg(
        f"v.harlan@{DOMAIN}", f"e.vasik@{DOMAIN}",
        "Locomotion milestone?",
        "Elena — where are we on the locomotion milestone? The principals are asking.\n\n— Viktor",
        "Mon, 14 Oct 2025 09:30:00 +0000",
        msg_id="<harlan-locomotion-001@boreas.local>"))

    save_eml("v.harlan", "03_holiday_schedule.eml", make_msg(
        f"hr@{DOMAIN}", f"all-staff@{DOMAIN}",
        "Holiday Schedule — December 2025",
        "Dear all,\n\nPlease note the following holiday schedule:\n"
        "- Dec 24-25: Christmas (office closed)\n"
        "- Dec 31 - Jan 1: New Year (office closed)\n"
        "- Essential personnel: contact your department head for coverage schedule.\n\n"
        "Happy holidays,\nHR Department",
        "Fri, 15 Nov 2025 14:00:00 +0000"))

    save_eml("v.harlan", "04_allhands_notes.eml", make_msg(
        f"v.harlan@{DOMAIN}", f"all-staff@{DOMAIN}",
        "All-Hands Meeting Notes — November",
        "Team,\n\nThank you for attending today's all-hands. Key points:\n\n"
        "1. Q3 results exceeded expectations — great work.\n"
        "2. Several new consulting contracts signed.\n"
        "3. Engineering team has made significant progress on internal initiatives.\n"
        "4. Reminder: information compartmentalization is critical. If you don't have "
        "a need to know, don't ask.\n\n"
        "Keep up the excellent work.\n\n— Viktor",
        "Wed, 05 Nov 2025 16:00:00 +0000"))


def vasik_inbox(status_report_path):
    """CTO inbox — technical emails, flag 8 in attachment."""
    # Reply to Harlan about locomotion
    save_eml("e.vasik", "01_re_locomotion.eml", make_msg(
        f"e.vasik@{DOMAIN}", f"v.harlan@{DOMAIN}",
        "Re: Locomotion milestone?",
        "Viktor,\n\nLocomotion is 100%. Weapons integration is on track. "
        "We are waiting on the primary power source.\n\n"
        "MIDNIGHT-7 exceeded all projections. Bipedal stability at 120m is achievable. "
        "The tail counterbalance responds within 18ms — well under our 20ms requirement.\n\n"
        "I've attached the full status report.\n\n— Elena",
        "Mon, 14 Oct 2025 10:45:00 +0000",
        in_reply_to="<harlan-locomotion-001@boreas.local>"))

    # Same email but with the PDF attachment (this is the one participants find)
    save_eml("e.vasik", "02_status_with_attachment.eml", make_msg_with_attachment(
        f"e.vasik@{DOMAIN}", f"v.harlan@{DOMAIN}",
        "Re: Locomotion milestone? [with attachment]",
        "Viktor,\n\nLocomotion is 100%. Weapons integration is on track. "
        "We are waiting on the primary power source.\n\n"
        "See attached status report for full details.\n\n— Elena",
        "Mon, 14 Oct 2025 10:47:00 +0000",
        status_report_path, "project_status_report_oct2025.pdf"))

    # Technical thread with lab team
    save_eml("e.vasik", "03_midnight7_results.eml", make_msg(
        f"e.vasik@{DOMAIN}", f"r.tanaka@{DOMAIN}",
        "Re: MIDNIGHT-7 Results",
        "Ryo,\n\nOutstanding work on MIDNIGHT-7. The bipedal stability numbers are "
        "exactly what we needed. 120m height with 18ms balance response — the tail "
        "counterbalance design is proven.\n\n"
        "Begin preparing for MIDNIGHT-8 (full integration with reactor power). "
        "We'll schedule this once the Novikov delivery is confirmed.\n\n"
        "Keep this between us until I brief Harlan.\n\n— Elena",
        "Tue, 29 Oct 2025 07:15:00 +0000",
        cc=f"p.nielsen@{DOMAIN}"))

    # Email to procurement about Kursk shipment
    save_eml("e.vasik", "04_kursk_expedite.eml", make_msg(
        f"e.vasik@{DOMAIN}", f"m.webb@{DOMAIN}",
        "URGENT: Expedite Kursk shipment",
        "Marcus,\n\nWe need expedited delivery of the Kursk shipment (PO-2847). "
        "We cannot slip the integration window. The actuators are the last "
        "mechanical component needed before reactor installation.\n\n"
        "Please contact Kursk Heavy Industries directly and authorize "
        "expedited shipping at whatever cost.\n\n— Elena",
        "Wed, 16 Jul 2025 08:30:00 +0000"))


def chen_inbox():
    """Terminated engineer inbox — the narrative thread."""
    save_eml("j.chen", "01_po_question.eml", make_msg(
        f"j.chen@{DOMAIN}", f"p.nielsen@{DOMAIN}",
        "Question about PO-2847",
        "Per,\n\nI was reviewing procurement docs for the quarterly audit and noticed "
        "PO-2847 — hydraulic actuators rated for 200 tons of force from Kursk Heavy "
        "Industries. $12 million.\n\n"
        "This doesn't match any client project I'm aware of. What is this for?\n\n"
        "— James",
        "Mon, 08 Sep 2025 14:22:00 +0000",
        msg_id="<chen-po-001@boreas.local>"))

    save_eml("j.chen", "02_manager_reply.eml", make_msg(
        f"p.nielsen@{DOMAIN}", f"j.chen@{DOMAIN}",
        "Re: Question about PO-2847",
        "James,\n\nThat's above your clearance. Focus on your assigned deliverables.\n\n"
        "— Per",
        "Mon, 08 Sep 2025 15:10:00 +0000",
        in_reply_to="<chen-po-001@boreas.local>"))

    save_eml("j.chen", "03_chen_followup.eml", make_msg(
        f"j.chen@{DOMAIN}", f"p.nielsen@{DOMAIN}",
        "Re: Re: Question about PO-2847",
        "Per,\n\nI pulled the specs on the actuators. These are rated for something "
        "enormous — 200 tons each, 48 units. Combined force capacity of 9,600 tons. "
        "That's not for any manufacturing process I've ever seen.\n\n"
        "Is this a weapons program? I signed an NDA but I didn't sign up for—\n\n"
        "— James",
        "Tue, 09 Sep 2025 09:45:00 +0000",
        in_reply_to="<chen-po-001@boreas.local>"))

    save_eml("j.chen", "04_termination.eml", make_msg(
        f"hr@{DOMAIN}", f"j.chen@{DOMAIN}",
        "Notice of Employment Termination — Effective Immediately",
        "Mr. Chen,\n\nThis email serves as formal notification that your employment "
        "with Boreas Systems Inc. is terminated effective immediately.\n\n"
        "Reason: Violation of Information Security Policy, Section 4.2 — "
        "Unauthorized Access to Classified Materials.\n\n"
        "Your access to all company systems has been revoked. Please return "
        "all company property to the front desk.\n\n"
        "Your final paycheck will be mailed within 14 business days.\n\n"
        "Human Resources Department\n"
        "Boreas Systems Inc.",
        "Fri, 12 Sep 2025 11:00:00 +0000"))


def kowalski_inbox():
    """IT admin inbox — flag 10 welcome email, creds backup in sent."""
    # Flag 10: Welcome email
    save_eml("d.kowalski", "01_welcome.eml", make_msg(
        f"postmaster@{DOMAIN}", f"d.kowalski@{DOMAIN}",
        "Welcome to your Boreas Systems mailbox",
        f"Welcome, Dariusz!\n\n"
        f"Your email account has been provisioned.\n\n"
        f"Username: d.kowalski\n"
        f"Server: mail.boreas.local\n"
        f"Webmail: http://mail.boreas.local\n\n"
        f"Account activation token: {FLAG_10}\n\n"
        f"Please change your password at your earliest convenience.\n\n"
        f"— Mail System",
        "Mon, 15 Jan 2024 08:00:00 +0000"))

    # Creds backup (in sent folder but we put it in inbox for simplicity)
    save_eml("d.kowalski", "02_creds_backup.eml", make_msg(
        f"d.kowalski@{DOMAIN}", f"d.kowalski@{DOMAIN}",
        "creds backup",
        "Backing up creds in case I forget:\n\n"
        "Wiki admin panel:\n"
        "  URL: http://intranet.boreas.local/admin\n"
        "  User: admin\n"
        "  Pass: admin\n\n"
        "File share service account:\n"
        "  Server: fileserv.boreas.local\n"
        "  User: svc-fileshare\n"
        "  Pass: F1l3Sh@r3Svc!\n"
        "  Shares: \\\\fileserv\\IT (has backup logs, network diagrams)\n\n"
        "SCADA gateway:\n"
        "  URL: http://scada-gw.boreas.local:8080\n"
        "  On VLAN 40 — not directly reachable from corporate\n\n"
        "Don't lose this.\n— D",
        "Wed, 05 Mar 2025 23:15:00 +0000"))

    # SCADA VLAN ticket
    save_eml("d.kowalski", "03_scada_vlan.eml", make_msg(
        f"d.kowalski@{DOMAIN}", f"e.vasik@{DOMAIN}",
        "Re: SCADA network isolation — DONE",
        "Elena,\n\nAs requested, I've isolated the generator controls on VLAN 40. "
        "Access via scada-gw.internal (10.10.40.10).\n\n"
        "The HMI web interface is on port 8080. Modbus PLC backend on port 502.\n\n"
        "I've set up the svc-scada service account for the web interface. "
        "Credentials are in the AD under ServiceAccounts OU.\n\n"
        "Only SCADA-Admins group has access. Currently that's you, me, and the "
        "svc-scada account.\n\n— Dariusz",
        "Thu, 20 Feb 2025 11:30:00 +0000"))


def morrison_inbox(rotation_path):
    """Security lead inbox — guard rotation, Petrov concerns."""
    # Guard rotation with attachment
    save_eml("s.morrison", "01_rotation_schedule.eml", make_msg_with_attachment(
        f"s.morrison@{DOMAIN}", f"security-team@{DOMAIN}",
        "Guard Rotation — November 2025",
        "Team,\n\nAttached is the November guard rotation schedule. "
        "Please review your assignments and confirm by Friday.\n\n"
        "Reminder: night shift guards must check in at all posts per the "
        "required patrol route. No exceptions.\n\n— Morrison",
        "Fri, 31 Oct 2025 16:00:00 +0000",
        rotation_path, "guard_rotation_november_2025.xlsx"))

    # Petrov concerns
    save_eml("s.morrison", "02_petrov_access.eml", make_msg(
        f"s.morrison@{DOMAIN}", f"v.harlan@{DOMAIN}",
        "SECURITY CONCERN: Guard Petrov — unusual access patterns",
        "Viktor,\n\nI'm flagging unusual badge access patterns for Guard Petrov.\n\n"
        "Over the past month, Petrov has accessed the underground hatch entrance "
        "6 times, all between 02:00-03:00 AM — outside his scheduled patrol route "
        "and shift times. Each time, his badge shows a gap of 40-50 minutes with "
        "no activity at any checkpoint, followed by an exit at the parking lot.\n\n"
        "No other guard has accessed that entrance during those hours.\n\n"
        "I've pulled the full badge logs and cross-referenced with the rotation "
        "schedule. His patrol assignments during those times were for Building A "
        "and the perimeter — nowhere near the hatch.\n\n"
        "Recommend we terminate Guard Petrov and review the hatch access controls.\n\n"
        "— Sarah Morrison\nSecurity Lead",
        "Mon, 03 Nov 2025 08:15:00 +0000"))

    # Follow-up
    save_eml("s.morrison", "03_petrov_followup.eml", make_msg(
        f"v.harlan@{DOMAIN}", f"s.morrison@{DOMAIN}",
        "Re: SECURITY CONCERN: Guard Petrov",
        "Sarah,\n\nThank you for flagging this. I'll discuss with Elena. "
        "In the meantime, do NOT confront Petrov directly. Continue monitoring.\n\n"
        "— Viktor",
        "Mon, 03 Nov 2025 09:45:00 +0000"))


def webb_inbox():
    """COO inbox — procurement and logistics."""
    save_eml("m.webb", "01_kursk_response.eml", make_msg(
        f"m.webb@{DOMAIN}", f"e.vasik@{DOMAIN}",
        "Re: URGENT: Expedite Kursk shipment",
        "Elena,\n\nI've contacted Kursk directly. They can expedite — adds $180K "
        "to shipping costs but gets the actuators here 2 weeks early.\n\n"
        "I've authorized the expedite. PO-2847 delivery now confirmed for "
        "August 15.\n\n— Marcus",
        "Wed, 16 Jul 2025 14:20:00 +0000"))

    save_eml("m.webb", "02_reactor_logistics.eml", make_msg(
        f"m.webb@{DOMAIN}", f"e.vasik@{DOMAIN}",
        "Reactor delivery logistics",
        "Elena,\n\nNovikov Energy Systems has confirmed the reactor delivery for "
        "November 25. They'll provide a 3-person installation team.\n\n"
        "Requirements:\n"
        "- Building B loading dock cleared for oversized delivery\n"
        "- Radiation monitoring team on standby\n"
        "- Underground access route must be clear for transport to installation site\n"
        "- Security perimeter during delivery (Morrison's team)\n\n"
        "I'll coordinate the logistics. Let me know if Engineering needs "
        "anything special for the installation.\n\n— Marcus",
        "Mon, 10 Nov 2025 10:00:00 +0000"))


# ============================================
# MAIN
# ============================================

def main():
    print("Building A1 Mail Server Content")
    print("=" * 50)
    make_dirs()

    print("\n--- Attachments ---")
    status_report_path = build_status_report_pdf()
    rotation_path = build_guard_rotation_xlsx()

    print("\n--- v.harlan inbox ---")
    harlan_inbox()

    print("\n--- e.vasik inbox ---")
    vasik_inbox(status_report_path)

    print("\n--- j.chen inbox ---")
    chen_inbox()

    print("\n--- d.kowalski inbox ---")
    kowalski_inbox()

    print("\n--- s.morrison inbox ---")
    morrison_inbox(rotation_path)

    print("\n--- m.webb inbox ---")
    webb_inbox()

    print("\n" + "=" * 50)
    total = sum(len(files) for _, _, files in os.walk(BASE))
    print(f"Total files: {total}")
    for root, dirs, files in sorted(os.walk(BASE)):
        rel = os.path.relpath(root, BASE)
        for f in sorted(files):
            size = os.path.getsize(os.path.join(root, f))
            print(f"  {rel}/{f} ({size:,} bytes)")


if __name__ == "__main__":
    main()
