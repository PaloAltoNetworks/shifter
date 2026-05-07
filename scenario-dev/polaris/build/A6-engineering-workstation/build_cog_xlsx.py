#!/usr/bin/env python3
"""
Build the real XLSX for A6 Nielsen's center_of_gravity_analysis.xlsx.
Three worksheets: Frame, Locomotion, Integration (hidden).
Flag 26 in a CONCATENATE-like formula on the Integration sheet.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, numbers

BASE = "/tmp/a6-content/home/p.nielsen/designs"

FLAG_26 = "FLAG{7e2b0c5d9a4f8163}"


def build():
    os.makedirs(BASE, exist_ok=True)
    path = os.path.join(BASE, "center_of_gravity_analysis.xlsx")

    wb = openpyxl.Workbook()
    hfill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")

    # === Sheet 1: Frame ===
    ws_frame = wb.active
    ws_frame.title = "Frame"
    for col, h in enumerate(["Component", "Height_m", "Mass_t", "COG_Height_m"], 1):
        c = ws_frame.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill

    frame_data = [
        ("Primary_frame", 120.4, 12000, 60.2),
        ("Dorsal_armor", 115, 3500, 57.5),
        ("Head_unit", 120.4, 800, 120.4),
        ("Internal_systems", 80, 2500, 80),
    ]
    for r, row in enumerate(frame_data, 2):
        for c, val in enumerate(row, 1):
            ws_frame.cell(row=r, column=c, value=val)

    # Add a sum row
    ws_frame.cell(row=7, column=1, value="TOTAL")
    ws_frame.cell(row=7, column=1).font = Font(bold=True)
    ws_frame.cell(row=7, column=3, value="=SUM(C2:C5)")
    ws_frame.cell(row=7, column=3).font = Font(bold=True)

    for col in range(1, 5):
        ws_frame.column_dimensions[chr(64 + col)].width = 20

    # === Sheet 2: Locomotion ===
    ws_loco = wb.create_sheet("Locomotion")
    for col, h in enumerate(["Component", "Height_m", "Mass_t", "COG_Height_m"], 1):
        c = ws_loco.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill

    loco_data = [
        ("Left_leg", 60, 24000, 30),
        ("Right_leg", 60, 24000, 30),
        ("Tail_base", 90, 1200, 90),
        ("Tail_segments_2_10", "varies", 7300, 55),
    ]
    for r, row in enumerate(loco_data, 2):
        for c, val in enumerate(row, 1):
            ws_loco.cell(row=r, column=c, value=val)

    ws_loco.cell(row=7, column=1, value="TOTAL")
    ws_loco.cell(row=7, column=1).font = Font(bold=True)
    ws_loco.cell(row=7, column=3, value="=SUM(C2:C5)")

    for col in range(1, 5):
        ws_loco.column_dimensions[chr(64 + col)].width = 22

    # === Sheet 3: Integration (HIDDEN) ===
    ws_int = wb.create_sheet("Integration")
    ws_int.sheet_state = "hidden"

    for col, h in enumerate(["Component", "Mass_t", "COG_m", "Moment"], 1):
        c = ws_int.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill

    int_data = [
        ("Total_frame", "=Frame!C7", 64.2, "=B2*C2"),
        ("Total_locomotion", "=Locomotion!C7", 37.1, "=B3*C3"),
        ("Total_arms", 4200, 92, "=B4*C4"),
        ("Total_weapons", 3800, 95, "=B5*C5"),
        ("TOTAL", "=SUM(B2:B5)", "=D6/B6", "=SUM(D2:D5)"),
    ]
    for r, row in enumerate(int_data, 2):
        for c, val in enumerate(row, 1):
            ws_int.cell(row=r, column=c, value=val)

    # Platform height
    ws_int.cell(row=8, column=1, value="Platform height (m)")
    ws_int.cell(row=8, column=2, value=120.4)

    # The flag — in a cell that combines data from all three sheets
    ws_int.cell(row=10, column=1, value="Integration verification code")
    ws_int.cell(row=10, column=1).font = Font(bold=True)
    ws_int.cell(row=10, column=2, value=FLAG_26)
    ws_int.cell(row=10, column=2).font = Font(color="FF0000", bold=True)

    # Add a note about what this sheet is
    ws_int.cell(row=12, column=1, value="This worksheet contains combined COG analysis")
    ws_int.cell(row=13, column=1, value="across all subsystems. Hidden per security policy.")
    ws_int.cell(row=14, column=1, value="Unhide this sheet to view integration calculations.")

    for col in range(1, 5):
        ws_int.column_dimensions[chr(64 + col)].width = 22

    wb.save(path)
    print(f"Created: {path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Integration sheet state: {ws_int.sheet_state}")

    # Verify the flag is in the hidden sheet
    wb2 = openpyxl.load_workbook(path)
    ws_check = wb2["Integration"]
    val = ws_check.cell(row=10, column=2).value
    print(f"  Flag 26 in Integration!B10: {val}")
    assert val == FLAG_26, f"Flag mismatch: {val}"
    print("  VERIFIED")


if __name__ == "__main__":
    build()
